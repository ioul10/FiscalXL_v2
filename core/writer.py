"""
core/writer.py
Construit l'Excel 5 feuilles à partir des données extraites.
Style propre, lisible, professionnel.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Palette de couleurs ───────────────────────────────────────────────────────
C_DARK   = "1F4E79"   # bleu foncé — titres principaux
C_MED    = "2E75B6"   # bleu moyen — headers colonnes
C_LIGHT  = "D6E4F0"   # bleu très clair — lignes calculées / totaux
C_STRIPE = "F2F7FC"   # bleu quasi-blanc — lignes alternées
C_WHITE  = "FFFFFF"
C_GOLD   = "FFF2CC"   # jaune — cellules saisie manuelle
C_BLACK  = "000000"
C_GRAY   = "666666"

NUM_FMT  = '#,##0.00'
NUM_FMT0 = '#,##0'

thin  = Side(style='thin',   color="CCCCCC")
thick = Side(style='medium', color=C_DARK)
BORDER_THIN  = Border(left=thin, right=thin, top=thin, bottom=thin)
BORDER_THICK = Border(bottom=Side(style='medium', color=C_DARK))


def _c(ws, row, col, value='', bg=C_WHITE, fg=C_BLACK, bold=False,
       align='left', sz=9, wrap=False, indent=0, num_fmt=None, border=None):
    """Écrire une cellule avec style complet."""
    cell = ws.cell(row, col)
    cell.value = value
    cell.font      = Font(name="Calibri", size=sz, bold=bold, color=fg)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center",
                               wrap_text=wrap, indent=indent)
    if num_fmt:
        cell.number_format = num_fmt
    if border:
        cell.border = border
    return cell


def _title_row(ws, row, text, n_cols=5, bg=C_DARK, fg=C_WHITE, sz=11):
    """Ligne de titre fusionnée."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    _c(ws, row, 1, text, bg=bg, fg=fg, bold=True, align='center', sz=sz)
    ws.row_dimensions[row].height = 22


def _header_row(ws, row, headers, widths, bg=C_MED, fg=C_WHITE):
    """Ligne d'en-têtes colonnes."""
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        _c(ws, row, ci, h, bg=bg, fg=fg, bold=True, align='center', sz=9, wrap=True)
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[row].height = 28


def _info_block(ws, row, info: dict, n_cols=5):
    """Bloc d'informations société sous le titre."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    _c(ws, row, 1, info.get('societe', ''), bg=C_STRIPE, bold=True, sz=10)
    _c(ws, row, 4, f"IF: {info.get('identifiant_fiscal', '')}", bg=C_STRIPE, sz=9, align='center')
    _c(ws, row, 5, info.get('exercice', ''), bg=C_STRIPE, sz=9, align='center', wrap=True)
    ws.row_dimensions[row].height = 18
    return row + 1


# ══════════════════════════════════════════════════════════════════════════════
# FEUILLE 1 — IDENTIFICATION
# ══════════════════════════════════════════════════════════════════════════════

def write_identification(wb, info: dict):
    ws = wb.create_sheet("1 - Identification")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 45

    r = 2
    _title_row(ws, r, "IDENTIFICATION FISCALE", n_cols=2, sz=12)
    r += 2

    fields = [
        ("Société / Raison sociale",   info.get('societe', '')),
        ("Identifiant Fiscal (IF)",    info.get('identifiant_fiscal', '')),
        ("Article IS",                 info.get('article_is', '')),
        ("Exercice comptable",         info.get('exercice', '')),
        ("Forme juridique",            info.get('forme_juridique', '')),
        ("Activité principale",        info.get('activite', '')),
        ("Siège social",               info.get('siege', '')),
    ]

    for i, (label, value) in enumerate(fields):
        bg = C_STRIPE if i % 2 == 0 else C_WHITE
        _c(ws, r, 1, label, bg=bg, bold=True, sz=10)
        _c(ws, r, 2, value, bg=bg, sz=10)
        ws.row_dimensions[r].height = 20
        r += 1


# ══════════════════════════════════════════════════════════════════════════════
# FEUILLE 2 — BILAN ACTIF
# ══════════════════════════════════════════════════════════════════════════════

def write_actif(wb, info: dict, rows: list):
    ws = wb.create_sheet("2 - Bilan Actif")
    ws.sheet_view.showGridLines = False

    r = 3
    _title_row(ws, r, "BILAN ACTIF", n_cols=5)
    r += 1
    _info_block(ws, r, info)
    r += 1

    hdrs   = ["DÉSIGNATION", "BRUT", "AMORT. & PROV.", "NET — EXERCICE N", "NET — EXERCICE N-1"]
    widths = [48, 18, 18, 18, 18]
    _header_row(ws, r, hdrs, widths)
    ws.freeze_panes = f'A{r+1}'
    r += 1

    # Mots-clés pour détecter les lignes de totaux / sections
    TOTAL_KW  = ['total', 'sous-total']
    SECTION_KW = ['immobilisations en non', 'immobilisations incorporelles',
                  'immobilisations corporelles', 'immobilisations financières',
                  'actif circulant', 'stocks', 'créances', 'trésorerie']

    for i, (label, vals) in enumerate(rows):
        ll = label.lower()
        is_total   = any(kw in ll for kw in TOTAL_KW)
        is_section = any(kw in ll for kw in SECTION_KW)

        if is_total:
            bg, fg, bold, ht = C_LIGHT, C_DARK, True, 16
        elif is_section:
            bg, fg, bold, ht = "E8F1FB", C_DARK, True, 15
        else:
            bg = C_STRIPE if i % 2 == 0 else C_WHITE
            fg, bold, ht = C_BLACK, False, 14

        _c(ws, r, 1, label, bg=bg, fg=fg, bold=bold, sz=9,
           indent=0 if (is_total or is_section) else 1)
        for ci, v in enumerate(vals[:4], 2):
            _c(ws, r, ci, v or 0, bg=bg, fg=fg, bold=bold,
               align='right', sz=9, num_fmt=NUM_FMT)
        ws.row_dimensions[r].height = ht
        r += 1

    return r


# ══════════════════════════════════════════════════════════════════════════════
# FEUILLE 3 — BILAN PASSIF
# ══════════════════════════════════════════════════════════════════════════════

def write_passif(wb, info: dict, rows: list):
    ws = wb.create_sheet("3 - Bilan Passif")
    ws.sheet_view.showGridLines = False

    r = 3
    _title_row(ws, r, "BILAN PASSIF", n_cols=3)
    r += 1
    _info_block(ws, r, info, n_cols=3)
    r += 1

    hdrs   = ["DÉSIGNATION", "EXERCICE N", "EXERCICE N-1"]
    widths = [52, 20, 20]
    _header_row(ws, r, hdrs, widths)
    ws.freeze_panes = f'A{r+1}'
    r += 1

    TOTAL_KW   = ['total', 'sous-total']
    SECTION_KW = ['capitaux propres', 'financement', 'dettes', 'trésorerie', 'provisions']

    for i, (label, vals) in enumerate(rows):
        ll = label.lower()
        is_total   = any(kw in ll for kw in TOTAL_KW)
        is_section = any(kw in ll for kw in SECTION_KW)

        if is_total:
            bg, fg, bold, ht = C_LIGHT, C_DARK, True, 16
        elif is_section:
            bg, fg, bold, ht = "E8F1FB", C_DARK, True, 15
        else:
            bg = C_STRIPE if i % 2 == 0 else C_WHITE
            fg, bold, ht = C_BLACK, False, 14

        _c(ws, r, 1, label, bg=bg, fg=fg, bold=bold, sz=9,
           indent=0 if (is_total or is_section) else 1)
        for ci, v in enumerate(vals[:2], 2):
            _c(ws, r, ci, v or 0, bg=bg, fg=fg, bold=bold,
               align='right', sz=9, num_fmt=NUM_FMT)
        ws.row_dimensions[r].height = ht
        r += 1


# ══════════════════════════════════════════════════════════════════════════════
# FEUILLE 4 — CPC
# ══════════════════════════════════════════════════════════════════════════════

def write_cpc(wb, info: dict, rows: list):
    ws = wb.create_sheet("4 - CPC")
    ws.sheet_view.showGridLines = False

    r = 3
    _title_row(ws, r, "COMPTE DE PRODUITS ET CHARGES (Hors Taxes)", n_cols=5)
    r += 1
    _info_block(ws, r, info)
    r += 1

    hdrs   = ["DÉSIGNATION", "PROPRES À\nL'EXERCICE",
              "EXERCICES\nPRÉCÉDENTS", "TOTAUX\nEXERCICE N", "TOTAUX\nEXERCICE N-1"]
    widths = [48, 18, 18, 18, 18]
    _header_row(ws, r, hdrs, widths)
    ws.freeze_panes = f'A{r+1}'
    r += 1

    TOTAL_KW   = ['total i', 'total ii', 'total iv', 'total v',
                  'total viii', 'total ix', 'total des produits', 'total des charges']
    RESULT_KW  = ['résultat', 'resultat']
    SECTION_KW = ['produits d\'exploitation', 'charges d\'exploitation',
                  'produits financiers', 'charges financières',
                  'produits non courants', 'charges non courantes']

    for i, (label, vals) in enumerate(rows):
        ll = label.lower()
        is_total   = any(kw in ll for kw in TOTAL_KW)
        is_result  = any(kw in ll for kw in RESULT_KW)
        is_section = any(kw in ll for kw in SECTION_KW)

        if is_total or is_result:
            bg, fg, bold, ht = C_LIGHT, C_DARK, True, 16
        elif is_section:
            bg, fg, bold, ht = "E8F1FB", C_DARK, True, 15
        else:
            bg = C_STRIPE if i % 2 == 0 else C_WHITE
            fg, bold, ht = C_BLACK, False, 14

        _c(ws, r, 1, label, bg=bg, fg=fg, bold=bold, sz=9,
           indent=0 if (is_total or is_result or is_section) else 1)
        for ci, v in enumerate(vals[:4], 2):
            _c(ws, r, ci, v or 0, bg=bg, fg=fg, bold=bold,
               align='right', sz=9, num_fmt=NUM_FMT)
        ws.row_dimensions[r].height = ht
        r += 1


# ══════════════════════════════════════════════════════════════════════════════
# FEUILLE 5 — ESG
# ══════════════════════════════════════════════════════════════════════════════

def write_esg(wb, info: dict, rows: list):
    ws = wb.create_sheet("5 - ESG")
    ws.sheet_view.showGridLines = False

    r = 3
    _title_row(ws, r, "ÉTAT DE SOLDES DE GESTION (E.S.G)", n_cols=3)
    r += 1
    _info_block(ws, r, info, n_cols=3)
    r += 1

    hdrs   = ["DÉSIGNATION", "EXERCICE N", "EXERCICE N-1"]
    widths = [52, 20, 20]
    _header_row(ws, r, hdrs, widths)
    ws.freeze_panes = f'A{r+1}'
    r += 1

    CALC_KW = ['marge brute', 'production de l', 'consommations de l',
               'valeur ajoutee', 'valeur ajoutée', 'excedent brut', 'insuffisance brute',
               'resultat d\'exploitation', 'résultat d\'exploitation',
               'resultat financier', 'résultat financier',
               'resultat courant', 'résultat courant',
               'resultat non courant', 'résultat non courant',
               'resultat net de l', 'résultat net de l',
               'capacite d\'autofinancement', 'capacité d\'autofinancement',
               'autofinancement']

    TFR_KW = ['benefice', 'bénéfice', 'perte']

    for i, (label, vals) in enumerate(rows):
        ll = label.lower()
        is_calc = any(kw in ll for kw in CALC_KW)
        is_tfr  = any(kw in ll for kw in TFR_KW)

        if is_calc:
            bg, fg, bold, ht = C_LIGHT, C_DARK, True, 16
        elif is_tfr:
            bg, fg, bold, ht = C_STRIPE, C_GRAY, False, 14
        else:
            bg = C_STRIPE if i % 2 == 0 else C_WHITE
            fg, bold, ht = C_BLACK, False, 14

        _c(ws, r, 1, label, bg=bg, fg=fg, bold=bold, sz=9,
           indent=0 if is_calc else 1)
        for ci, v in enumerate(vals[:2], 2):
            _c(ws, r, ci, v or 0, bg=bg, fg=fg, bold=bold,
               align='right', sz=9, num_fmt=NUM_FMT)
        ws.row_dimensions[r].height = ht
        r += 1

    # Ligne distributions de bénéfices (saisie manuelle)
    r += 1
    _c(ws, r, 1, "Distributions de bénéfices", bg=C_GOLD, sz=9, indent=1)
    for ci in [2, 3]:
        c = ws.cell(r, ci)
        c.value = 0
        c.fill  = PatternFill("solid", fgColor=C_GOLD)
        c.font  = Font(name="Calibri", size=9, italic=True, color=C_GRAY)
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.number_format = NUM_FMT
    ws.row_dimensions[r].height = 14

    # Note en bas
    r += 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    note = ws.cell(r, 1)
    note.value = "(1) À l'exclusion des dotations relatives aux actifs et passifs circulants. " \
                 "(2) À l'exclusion des reprises relatives aux actifs circulants. " \
                 "(*) Cellule jaune = saisie manuelle."
    note.font  = Font(name="Calibri", size=8, italic=True, color=C_GRAY)
    note.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[r].height = 30


# ══════════════════════════════════════════════════════════════════════════════
# POINT D'ENTRÉE
# ══════════════════════════════════════════════════════════════════════════════

def build_excel(data: dict, output_path: str):
    """
    Construit l'Excel complet depuis les données extraites.
    data = {
        'info':   dict identification,
        'actif':  [(label, [brut,amort,netN,netN1]), ...],
        'passif': [(label, [exN, exN1]), ...],
        'cpc':    [(label, [propN, exercPrec, totN, totN1]), ...],
        'esg':    [(label, [exN, exN1]), ...],   # peut être vide
    }
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # supprimer la feuille par défaut

    info = data.get('info', {})

    write_identification(wb, info)
    write_actif(wb,  info, data.get('actif',  []))
    write_passif(wb, info, data.get('passif', []))
    write_cpc(wb,    info, data.get('cpc',    []))
    write_esg(wb,    info, data.get('esg',    []))

    wb.save(output_path)
    return output_path
